"""
Builds `Projects.xlsx` — the intake sheet the team fills in and the dashboard reads.

    python build_template.py

It is one file doing two jobs: the form people fill in, and the app's data source. Run it
once to create the sheet; it seeds itself from the old `My Tasks_Projects.xlsx` so the 27
existing projects are carried over rather than retyped.

Re-running is safe but DESTRUCTIVE: it rewrites the file from the old tracker and would
discard anything typed into `Projects.xlsx` since. It refuses to overwrite unless you
pass --force.

What is deliberately NOT migrated: "How It Was Built" and "Why It Was Built" are left
empty for every project. The old tracker never had those columns, and inventing them is
exactly what we must not do — they are the whole point of the new sheet.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).parent
OUT = HERE / "Projects.xlsx"
LEGACY = Path.home() / "Downloads" / "My Tasks_Projects.xlsx"

DATA_SHEET = "Projects"
GUIDE_SHEET = "Read me first"
LISTS_SHEET = "Lists"
LAST_ROW = 400       # validation + formatting applied this far down, ready to type into
LIST_RANGE_END = 60  # dropdowns read Lists rows 2..60, so new options just work

# The app enforces these five - its Active / Completed / Needs attention grouping keys
# off them, so this dropdown is validated strictly.
STATUSES = ["Not started", "In Progress", "Continuous", "Blocked", "On Hold", "Completed"]
PRIORITIES = ["P0 — Critical", "P1 — High", "P2 — Medium", "P3 — Long-Term"]
VERTICALS = [
    "Marketing", "Product", "Finance", "Operations", "OPM", "L & D",
    "Data / Analytics", "Leadership", "DRS", "Re-Commerce", "Tech", "People",
]

# (header, width, wrap, required, tooltip)
COLUMNS: list[tuple[str, int, bool, bool, str]] = [
    ("ID", 6, False, False,
     "Any unique number. Leave blank and one will be assigned by position."),
    ("Project Name", 38, True, True,
     "What the project is called. This is the headline on the dashboard, so make it "
     "readable — 'Leadership view of the central dashboard', not 'LV dash v2'."),
    ("Vertical", 16, False, True,
     "Which vertical this belongs to. Pick from the list, or type a new one if none fit."),
    ("Status", 14, False, True,
     "Pick from the dropdown, or type your own. Each status becomes its own filter on "
     "the dashboard; add one to the Lists tab to share it with everyone. The order of "
     "the Status column on the Lists tab sets the order of the filters."),
    ("Priority", 15, False, False, "Pick from the list."),
    ("Point of Contact", 20, True, True,
     "The ONE person to talk to about this project. Shown on every card. Use a real "
     "name, spelled consistently — separate multiple people with commas."),
    ("Contributors", 20, True, False,
     "Anyone else involved, comma separated. These appear under 'Built with'."),
    ("What It Is About", 60, True, True,
     "Two or three sentences a newcomer could understand: what this project is and what "
     "it produced. This is the first thing anyone reads on the dashboard."),
    ("How It Was Built", 60, True, False,
     "How it was actually put together — the tool or stack (Apps Script, Sheet, deck), "
     "where the data comes from, the approach, and any decision worth knowing about."),
    ("Why It Was Built", 60, True, False,
     "Why this was picked up at all — the problem, the request, the decision it had to "
     "support. NOT what changed afterwards; that goes in Impact. The old tracker had no "
     "column for this, which is why most projects cannot answer it."),
    ("Impact / What Changed", 50, True, False,
     "What was different afterwards. Numbers if you have them."),
    ("Time Saved", 16, True, False, "e.g. '2 hours a week'. Leave blank if unknown."),
    ("Money Saved", 16, True, False, "e.g. 'Rs 40,000/month'. Leave blank if unknown."),
    ("Start Date", 13, False, False, "When work began. Format dd-mm-yyyy."),
    ("Completed Date", 14, False, False,
     "Leave blank while the project is still open."),
    ("Steps / Checklist", 46, True, False,
     "One step per line, each starting with [x] for done or [ ] for outstanding. "
     "Example:\n[x] Collected the data\n[ ] Waiting on Marketing\n"
     "The dashboard turns this into a progress bar and a checklist."),
    ("Links", 40, True, False,
     "One URL per line — the dashboard reads each link's address to work out what the "
     "project was built as (Apps Script, Sheet, Doc, deck)."),
    ("Notes", 34, True, False, "Anything else worth recording."),
]

INK = "1D1D1F"
GREY = "F5F5F7"
REQ = "FFF8E7"          # faint warm tint marking required columns
HAIRLINE = "D2D2D7"

GUIDE = [
    ("Strategy Team — Project Registry", "title"),
    ("How to add your project", "h1"),
    ("", ""),
    ("Open the 'Projects' sheet and add one row per project. Fill the columns "
     "left to right; hover any heading for a fuller explanation of what belongs in it.",
     "body"),
    ("", ""),
    ("The four that matter most", "h2"),
    ("Point of Contact — the one person someone should ask about this project.", "bullet"),
    ("What It Is About — two or three sentences a newcomer could follow.", "bullet"),
    ("How It Was Built — the tool, the data source, the approach.", "bullet"),
    ("Why It Was Built — the reason it was picked up in the first place.", "bullet"),
    ("", ""),
    ("'Why' is not the same as 'Impact'. Why is the problem that caused the work to "
     "start; Impact is what was different once it shipped. The old tracker only had "
     "Impact, which is why most existing projects show 'no stated rationale' on the "
     "dashboard.", "body"),
    ("", ""),
    ("Nothing here is fixed", "h2"),
    ("This sheet belongs to the team, not to the tool. You can add or delete rows, add "
     "new verticals or statuses on the 'Lists' tab, reorder them, or add your own "
     "columns — the dashboard follows whatever the sheet says.", "body"),
    ("Add a vertical or a status: put it on the 'Lists' tab and it becomes a dropdown "
     "option and a filter.", "bullet"),
    ("Add a column: just add it. It shows up on each project's page under 'Also "
     "recorded'. No one needs to change any code.", "bullet"),
    ("Delete a column you don't want: that section stops appearing.", "bullet"),
    ("The one column that must stay is 'Project Name'.", "bullet"),
    ("", ""),
    ("Rules", "h2"),
    ("Status, Priority and Vertical have dropdowns so spellings stay consistent — but "
     "none of them block you. Type a new value whenever you need one, and add it to the "
     "'Lists' tab so everyone else gets it in their dropdown too.", "bullet"),
    ("Leave anything you genuinely do not know BLANK. The dashboard shows an honest "
     "'not recorded yet' note for empty fields. Please do not guess or pad — a blank is "
     "more useful than a made-up sentence.", "bullet"),
    ("Steps / Checklist: one per line, each starting with [x] done or [ ] outstanding.",
     "bullet"),
    ("Links: one URL per line.", "bullet"),
    ("", ""),
    ("Then what?", "h2"),
    ("Save the file. In the dashboard, open 'Data source' at the foot of the page and "
     "press 'Refresh from Excel' — your project appears immediately. Nothing needs "
     "rebuilding or redeploying.", "body"),
    ("", ""),
    ("Existing projects have been carried over from the old tracker. Their 'How It Was "
     "Built' and 'Why It Was Built' cells were left empty on purpose — the old sheet "
     "never captured them, and they were not invented. If a project is yours, please "
     "fill those two in.", "body"),
]


# ---------------------------------------------------------------------------
# Seed rows from the legacy tracker
# ---------------------------------------------------------------------------
def _text(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    return "" if s.lower() in {"nan", "none", "nat"} else s


def _key(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _text(name).lower()).strip()


def _urls(*cells) -> str:
    found, seen = [], set()
    for c in cells:
        for u in re.findall(r"https?://[^\s,;)\]]+", _text(c)):
            u = u.rstrip(".,;")
            if u not in seen:
                seen.add(u)
                found.append(u)
    return "\n".join(found)


def _strip_urls(v) -> str:
    return re.sub(r"\s*https?://[^\s,;)\]]+\s*,?", " ", _text(v)).strip(" ,")


def _date(v):
    if not _text(v):
        return None
    ts = pd.to_datetime(v, errors="coerce")
    return None if pd.isna(ts) else ts.to_pydatetime()


def seed_rows(legacy: Path) -> list[dict]:
    """Carry the old tracker's 27 projects across, without inventing the new fields."""
    if not legacy.exists():
        print(f"  ! {legacy.name} not found — writing an empty template.")
        return []

    sheets = pd.read_excel(legacy, sheet_name=None, dtype=object)
    master = sheets.get("All", pd.DataFrame()).dropna(how="all")
    detail = sheets.get("Completed", pd.DataFrame()).dropna(how="all")
    by_key = {_key(r.get("Task")): r.to_dict() for _, r in detail.iterrows()
              if _key(r.get("Task"))}

    rows = []
    for _, m in master.iterrows():
        m = m.to_dict()
        name = _text(m.get("Task"))
        if not name:
            continue
        d = by_key.get(_key(name), {})
        raw_id = _text(m.get("ID"))
        rows.append({
            "ID": int(float(raw_id)) if raw_id.replace(".", "").isdigit() else None,
            "Project Name": name,
            "Vertical": _text(m.get("Vertical")),
            "Status": _text(m.get("Status")),
            "Priority": _text(m.get("Priority")),
            "Point of Contact": _text(m.get("Answerable To")),
            "Contributors": _text(m.get("Dependencies")),
            # `Process Done` is the old sheet's "what was done" column.
            "What It Is About": _strip_urls(d.get("Process Done")),
            "How It Was Built": "",   # never captured — must not be invented
            "Why It Was Built": "",   # never captured — must not be invented
            "Impact / What Changed": _text(d.get("Impact")),
            "Time Saved": _text(d.get("Time Saved")),
            "Money Saved": _text(d.get("Money Saved")),
            "Start Date": _date(d.get("Created")),
            "Completed Date": _date(d.get("Completed")),
            "Steps / Checklist": _text(m.get("Subtasks")) or _text(d.get("Subtasks")),
            "Links": _urls(m.get("Notes"), d.get("Notes"), d.get("Process Done")),
            "Notes": _strip_urls(m.get("Notes")) or _strip_urls(d.get("Notes")),
        })
    print(f"  + carried over {len(rows)} projects from {legacy.name}")
    return rows


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------
def build(rows: list[dict]) -> Workbook:
    wb = Workbook()

    # ---- guide -----------------------------------------------------------
    g = wb.active
    g.title = GUIDE_SHEET
    g.sheet_view.showGridLines = False
    g.column_dimensions["A"].width = 3
    g.column_dimensions["B"].width = 104
    styles = {
        "title": (Font(size=20, bold=True, color=INK), 34),
        "h1": (Font(size=13, bold=True, color="6E6E73"), 22),
        "h2": (Font(size=12, bold=True, color=INK), 30),
        "body": (Font(size=11, color="424245"), None),
        "bullet": (Font(size=11, color="424245"), None),
        "": (Font(size=11), 10),
    }
    # The per-column guidance is ALSO written here, not only as header comments on the
    # Projects sheet. Excel cell comments do not reliably survive conversion to Google
    # Sheets, and this sheet is meant to be uploaded there — so the explanations have to
    # live somewhere that always makes the trip.
    guide_rows = list(GUIDE) + [("", ""), ("What each column means", "h2")]
    guide_rows += [
        (f"{header} — {'REQUIRED. ' if required else ''}{tip}", "coldef")
        for header, _w, _wrap, required, tip in COLUMNS
    ]

    styles["coldef"] = (Font(size=10, color="424245"), None)
    for i, (text, kind) in enumerate(guide_rows, start=2):
        cell = g.cell(row=i, column=2, value=("•  " + text) if kind == "bullet" else text)
        font, height = styles[kind]
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if height:
            g.row_dimensions[i].height = height
        elif kind in {"body", "bullet", "coldef"}:
            g.row_dimensions[i].height = max(16, 14 * (len(text) // 100 + 1))

    # ---- lists (dropdown sources) ---------------------------------------
    ls = wb.create_sheet(LISTS_SHEET)
    ls["A1"], ls["B1"], ls["C1"] = "Status", "Priority", "Vertical"
    for col, values in enumerate([STATUSES, PRIORITIES, sorted(VERTICALS)], start=1):
        ls.cell(row=1, column=col).font = Font(bold=True)
        for r, v in enumerate(values, start=2):
            ls.cell(row=r, column=col, value=v)
    ls.column_dimensions["A"].width = 16
    ls.column_dimensions["B"].width = 16
    ls.column_dimensions["C"].width = 20
    ls["E1"] = "These feed the dropdowns on the Projects sheet."
    ls["E2"] = ("Add a row to any column here and it becomes a new option straight away — "
                "no permission needed, nothing to rebuild.")
    ls["E3"] = ("The Status column also sets the ORDER of the filters on the dashboard. "
                "Reorder these rows and the dashboard reorders with them.")
    ls["E4"] = ("You can delete options too. Just keep the three headings in row 1 as "
                "they are, so the dropdowns keep finding them.")
    ls["E1"].font = Font(bold=True, size=11, color="1D1D1F")
    for ref in ("E2", "E3", "E4"):
        ls[ref].font = Font(size=10, color="6E6E73")
        ls[ref].alignment = Alignment(wrap_text=True, vertical="top")
    ls.column_dimensions["E"].width = 88
    for r in (2, 3, 4):
        ls.row_dimensions[r].height = 28

    # ---- projects --------------------------------------------------------
    ws = wb.create_sheet(DATA_SHEET, 1)
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color=HAIRLINE)

    for i, (header, width, wrap, required, tip) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=header)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=INK)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = Border(bottom=thin)
        note = ("REQUIRED. " if required else "Optional. ") + tip
        c.comment = Comment(note, "Registry", width=330, height=170)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 34

    headers = [c[0] for c in COLUMNS]
    for r, row in enumerate(rows, start=2):
        for i, header in enumerate(headers, start=1):
            ws.cell(row=r, column=i, value=row.get(header) or None)

    # formatting + validation down to LAST_ROW so typing into blank rows just works
    for i, (header, _w, wrap, required, _t) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(i)
        for r in range(2, LAST_ROW + 1):
            cell = ws.cell(row=r, column=i)
            cell.alignment = Alignment(
                wrap_text=wrap, vertical="top",
                horizontal="left" if i != 1 else "center",
            )
            cell.font = Font(size=10)
            cell.border = Border(bottom=Side(style="hair", color="EAEAEF"))
            if required:
                cell.fill = PatternFill("solid", fgColor=REQ)
        if header in {"Start Date", "Completed Date"}:
            for r in range(2, LAST_ROW + 1):
                ws.cell(row=r, column=i).number_format = "DD-MM-YYYY"

    # Every dropdown is a SUGGESTION, never a constraint:
    #   * showErrorMessage=False, so a value typed in by hand is always accepted;
    #   * the source range runs well past the current options, so adding a row to the
    #     Lists tab immediately makes that option appear in the dropdown.
    # An earlier version pointed at exactly the current options with strict enforcement,
    # which silently blocked the team from inventing a new status. Don't reintroduce that.
    def add_dv(col_header: str, list_col: str, msg: str) -> None:
        i = headers.index(col_header) + 1
        letter = get_column_letter(i)
        dv = DataValidation(
            type="list",
            formula1=f"={LISTS_SHEET}!${list_col}$2:${list_col}${LIST_RANGE_END}",
            allow_blank=True, showErrorMessage=False, showInputMessage=True,
        )
        dv.promptTitle, dv.prompt = col_header, msg
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{LAST_ROW}")

    add_dv("Status", "A",
           "Pick a status, or type a new one. To add it to this dropdown for everyone, "
           "put it on the Lists tab.")
    add_dv("Priority", "B",
           "Pick a priority, or type a new one. Add it to the Lists tab to share it.")
    add_dv("Vertical", "C",
           "Pick a vertical, or type a new one. Add it to the Lists tab to share it.")

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{LAST_ROW}"

    wb.active = 0
    return wb


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--force", action="store_true",
                    help="overwrite an existing Projects.xlsx (discards anything typed in it)")
    ap.add_argument("--legacy", default=str(LEGACY), help="old tracker to seed from")
    ap.add_argument("--out", default=str(OUT),
                    help="write somewhere else (useful when the file is open in Excel)")
    args = ap.parse_args()

    out = Path(args.out)
    if out.exists() and not args.force:
        print(f"Refusing to overwrite {out.name} — it may already have projects typed in.")
        print("Re-run with --force if you really want to rebuild it from the old tracker.")
        return 1

    print(f"Building {out.name}")
    wb = build(seed_rows(Path(args.legacy)))
    try:
        wb.save(out)
    except PermissionError:
        print(f"  ! {out.name} is locked — it is probably open in Excel.")
        print("    Close it and re-run, or pass --out to write elsewhere.")
        return 1
    print(f"  + wrote {out}")
    print(f"  + sheets: {GUIDE_SHEET} / {DATA_SHEET} / {LISTS_SHEET}")
    print("\nShare this file with the team. Point the dashboard's 'Data source' at it.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
